from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
from openpyxl import load_workbook
import arabic_reshaper
from bidi.algorithm import get_display



#-----------------------------------------
#Change these variables to control the font and the text size.
theFont = 'Arial Unicode'
theTextSize = 35
theTextWidth = 1
#Change these variables to control the color of the text in RGB values.
theRed = 249
theGreen = 239
theBlue = 89
#Change this variable to control the veritcal position of the text. use as small sample to test the position to find the best position.
thePosition = 245
#-----------------------------------------



# Load the workbook. change the name here if your xlsx file is named diffrently.
wb = load_workbook('template.xlsx')

# Get the active sheet
ws = wb.active

# Create a list of the names
names = [cell.value for cell in ws['A']]
files = [cell.value for cell in ws['C']]

# Load the image
img = Image.open('cert.jpg')
W,H = img.size

# Iterate through the names
for name, filename in zip(names, files):
    

    name = arabic_reshaper.reshape(name)
    name = get_display(name)
    # Generate a new image
    new_img = img.copy()
    W,H = new_img.size

    # Write the name on the image
    
    draw = ImageDraw.Draw(new_img)
    font = ImageFont.truetype(theFont, theTextSize)
    _, _, w, h = draw.textbbox((0, 0), name, font=font)
    draw.text(((W-w)/2, thePosition), name, font=font, stroke_width=theTextWidth, fill=(theRed, theGreen, theBlue))

    # Save the new image as a PDF.
    new_img.save(f'Result/{filename}.pdf')